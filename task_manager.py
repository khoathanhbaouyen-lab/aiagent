"""
Task Manager Module - Enhanced với Priority, Tags, Assignment
"""
import sqlite3
import os
import json
from typing import List, Optional, Dict
from datetime import datetime

USERS_DB_FILE = os.path.join("user_data", "users.sqlite")

def _get_db_conn():
    """Kết nối database"""
    return sqlite3.connect(USERS_DB_FILE, check_same_thread=False)

def get_all_users() -> List[Dict]:
    """Lấy danh sách tất cả users (email + name)"""
    conn = _get_db_conn()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT email, name, is_active 
        FROM users 
        WHERE is_active = 1
        ORDER BY name, email
    """)
    
    users = []
    for row in cursor.fetchall():
        users.append({
            'email': row['email'],
            'name': row['name'] or row['email'].split('@')[0]
        })
    
    conn.close()
    return users

# ============= SCHEMA MIGRATION =============
def migrate_task_schema():
    """Thêm các cột mới vào bảng user_tasks nếu chưa có"""
    conn = _get_db_conn()
    cursor = conn.cursor()
    
    # Lấy thông tin schema hiện tại
    cursor.execute("PRAGMA table_info(user_tasks);")
    columns = [col[1] for col in cursor.fetchall()]
    
    # Thêm các cột mới
    if 'priority' not in columns:
        cursor.execute("ALTER TABLE user_tasks ADD COLUMN priority TEXT DEFAULT 'medium'")
        print("✅ Đã thêm cột 'priority'")
    
    if 'tags' not in columns:
        cursor.execute("ALTER TABLE user_tasks ADD COLUMN tags TEXT")
        print("✅ Đã thêm cột 'tags'")
    
    if 'assigned_to' not in columns:
        cursor.execute("ALTER TABLE user_tasks ADD COLUMN assigned_to TEXT")
        print("✅ Đã thêm cột 'assigned_to'")
    
    if 'assigned_by' not in columns:
        cursor.execute("ALTER TABLE user_tasks ADD COLUMN assigned_by TEXT")
        print("✅ Đã thêm cột 'assigned_by'")
    
    conn.commit()
    conn.close()

# ============= CRUD FUNCTIONS =============
def create_task(
    user_email: str,
    title: str,
    description: Optional[str] = None,
    due_date: datetime = None,
    priority: str = "medium",
    tags: List[str] = None,
    assigned_to: Optional[str] = None,
    assigned_by: Optional[str] = None,
    recurrence_rule: Optional[str] = None
) -> int:
    """Tạo task mới với đầy đủ tính năng"""
    conn = _get_db_conn()
    cursor = conn.cursor()
    
    tags_json = json.dumps(tags) if tags else None
    
    cursor.execute("""
        INSERT INTO user_tasks 
        (user_email, title, description, due_date, priority, tags, 
         assigned_to, assigned_by, recurrence_rule, is_completed)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
    """, (user_email.lower(), title, description, due_date, priority, 
          tags_json, assigned_to, assigned_by, recurrence_rule))
    
    task_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    print(f"✅ [TaskManager] Created task #{task_id}: {title}")
    return task_id

def get_tasks(
    user_email: Optional[str] = None,
    status: str = "all",
    priority: Optional[str] = None,
    tags: Optional[List[str]] = None,
    assigned_to: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None
) -> List[Dict]:
    """Lấy danh sách tasks với filters nâng cao"""
    conn = _get_db_conn()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    query = "SELECT * FROM user_tasks WHERE 1=1"
    params = []
    
    if user_email:
        query += " AND (user_email = ? OR assigned_to = ?)"
        params.extend([user_email.lower(), user_email.lower()])
    
    if status == "uncompleted":
        query += " AND is_completed = 0"
    elif status == "completed":
        query += " AND is_completed = 1"
    
    if priority:
        query += " AND priority = ?"
        params.append(priority)
    
    if assigned_to:
        query += " AND assigned_to = ?"
        params.append(assigned_to.lower())
    
    if start_date:
        query += " AND due_date >= ?"
        params.append(start_date)
    
    if end_date:
        query += " AND due_date <= ?"
        params.append(end_date)
    
    # Filter by tags (JSON search)
    if tags:
        for tag in tags:
            query += " AND tags LIKE ?"
            params.append(f"%{tag}%")
    
    query += " ORDER BY priority DESC, due_date ASC"
    
    cursor.execute(query, params)
    tasks = []
    for row in cursor.fetchall():
        task = dict(row)
        # Parse tags JSON
        if task['tags']:
            try:
                task['tags'] = json.loads(task['tags'])
            except:
                task['tags'] = []
        else:
            task['tags'] = []
        tasks.append(task)
    
    conn.close()
    return tasks

def update_task(
    task_id: int,
    title: Optional[str] = None,
    description: Optional[str] = None,
    due_date: Optional[datetime] = None,
    priority: Optional[str] = None,
    tags: Optional[List[str]] = None,
    assigned_to: Optional[str] = None,
    recurrence_rule: Optional[str] = None
) -> bool:
    """Cập nhật task"""
    conn = _get_db_conn()
    cursor = conn.cursor()
    
    updates = []
    params = []
    
    if title:
        updates.append("title = ?")
        params.append(title)
    if description is not None:
        updates.append("description = ?")
        params.append(description)
    if due_date:
        updates.append("due_date = ?")
        params.append(due_date)
    if priority:
        updates.append("priority = ?")
        params.append(priority)
    if tags is not None:
        updates.append("tags = ?")
        params.append(json.dumps(tags))
    if assigned_to is not None:
        updates.append("assigned_to = ?")
        params.append(assigned_to.lower() if assigned_to else None)
    if recurrence_rule is not None:
        updates.append("recurrence_rule = ?")
        params.append(recurrence_rule)
    
    if not updates:
        conn.close()
        return False
    
    query = f"UPDATE user_tasks SET {', '.join(updates)} WHERE id = ?"
    params.append(task_id)
    
    cursor.execute(query, params)
    updated = cursor.rowcount > 0
    conn.commit()
    conn.close()
    
    return updated

def mark_complete(task_id: int, user_email: str) -> bool:
    """Đánh dấu hoàn thành"""
    conn = _get_db_conn()
    cursor = conn.cursor()
    
    cursor.execute(
        "UPDATE user_tasks SET is_completed = 1 WHERE id = ? AND (user_email = ? OR assigned_to = ?)",
        (task_id, user_email.lower(), user_email.lower())
    )
    
    updated = cursor.rowcount > 0
    conn.commit()
    conn.close()
    
    return updated

def delete_task(task_id: int, user_email: str) -> bool:
    """Xóa task"""
    conn = _get_db_conn()
    cursor = conn.cursor()
    
    cursor.execute(
        "DELETE FROM user_tasks WHERE id = ? AND (user_email = ? OR assigned_by = ?)",
        (task_id, user_email.lower(), user_email.lower())
    )
    
    deleted = cursor.rowcount > 0
    conn.commit()
    conn.close()
    
    return deleted

# ============= EXPORT FUNCTIONS =============
def export_tasks_csv(
    user_email: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None
) -> str:
    """Export tasks ra CSV"""
    import csv
    from io import StringIO
    
    tasks = get_tasks(
        user_email=user_email,
        start_date=start_date,
        end_date=end_date,
        status="all"
    )
    
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        'id', 'title', 'description', 'due_date', 'priority', 
        'tags', 'assigned_to', 'assigned_by', 'is_completed', 'created_at'
    ])
    
    writer.writeheader()
    for task in tasks:
        task['tags'] = ', '.join(task['tags']) if task['tags'] else ''
        writer.writerow(task)
    
    return output.getvalue()

def export_tasks_excel(
    user_email: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    output_path: str = "tasks_report.xlsx"
):
    """Export tasks ra Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("❌ Cần cài openpyxl: pip install openpyxl")
        return None
    
    tasks = get_tasks(
        user_email=user_email,
        start_date=start_date,
        end_date=end_date,
        status="all"
    )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks Report"
    
    # Header
    headers = ['ID', 'Title', 'Description', 'Due Date', 'Priority', 
               'Tags', 'Assigned To', 'Status', 'Created At']
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Data rows
    for task in tasks:
        ws.append([
            task['id'],
            task['title'],
            task['description'] or '',
            task['due_date'],
            task['priority'],
            ', '.join(task['tags']) if task['tags'] else '',
            task['assigned_to'] or '',
            'Completed' if task['is_completed'] else 'Pending',
            task['created_at']
        ])
    
    wb.save(output_path)
    print(f"✅ Exported to {output_path}")
    return output_path

# ============= STATISTICS =============
def get_task_stats(user_email: str) -> Dict:
    """Thống kê tasks"""
    conn = _get_db_conn()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN is_completed = 1 THEN 1 ELSE 0 END) as completed,
            SUM(CASE WHEN is_completed = 0 THEN 1 ELSE 0 END) as pending,
            SUM(CASE WHEN priority = 'high' AND is_completed = 0 THEN 1 ELSE 0 END) as high_priority
        FROM user_tasks
        WHERE user_email = ? OR assigned_to = ?
    """, (user_email.lower(), user_email.lower()))
    
    row = cursor.fetchone()
    conn.close()
    
    return {
        'total': row[0] or 0,
        'completed': row[1] or 0,
        'pending': row[2] or 0,
        'high_priority': row[3] or 0
    }

# Initialize on import
if __name__ == "__main__":
    migrate_task_schema()
